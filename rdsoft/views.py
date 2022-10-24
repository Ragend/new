from django.shortcuts import render
import openpyxl as op
from openpyxl.styles import PatternFill, Alignment
import time

# Create your views here.
def home(request):

    if len(request.FILES.getlist('document')) != 0:
        print(request.FILES.getlist('document'))
        if request.method == 'POST':
            files = request.FILES.getlist('document')
            for rdfile in files:

                print("file is...",rdfile)
                print("list is..",files)
                

                
            
                file = op.load_workbook(rdfile)
                sheets = file.sheetnames
                sheet = file[sheets[0]]
                sheet.merge_cells('N1:O1')
                sheet.cell(row=1, column=14).value="Valid upto : 12/12/21 \nPan  No: AYZPB638IJ"


                sheet._images
                del sheet._images[0]
                d = op.drawing.image.Image('C:\\Users\\MY BOOK\\Downloads\\rdimage.jpg')
                sheet.add_image(d,'A1')
                sheet.row_dimensions[1].height=55.20
                for i in range(2,13):

                    sheet.row_dimensions[i].height = 13.20
                    sheet.column_dimensions['M'].width = 18.56
                    sheet.column_dimensions['E'].width = 13.44
                    sheet.column_dimensions['F'].width = 14.89
                    sheet.column_dimensions['G'].width = 29.89
                    sheet.column_dimensions['L'].width = 14.00
                    sheet.column_dimensions['O'].width = 12.22
                    sheet.column_dimensions['B'].width = 3.11

                for i in range(14,sheet.max_row-1):
                    sheet.row_dimensions[i].height=32.00

                maxrow = sheet.max_row

                kkk = "J"+str(maxrow-1)+":Q"+str(maxrow-1)
                lll = "'Q"+str(maxrow-1)+"'"
                mmm = "J"+str(maxrow-1)+":O"+str(maxrow-1)

                print(kkk)
                print(mmm)
                sheet.unmerge_cells(kkk)

                sheet.delete_cols(16, 25)

                sheet.merge_cells(mmm)
                print("testing..... ", sheet['D14'].value)
                rupees = sheet.cell(row=maxrow, column=10).value
                #rupees = rupees.replace(',', '')

                sheet.merge_cells('B2:S2')
                sheet.unmerge_cells('B2:S2')
                sheet.merge_cells('B2:O2')

                sheet.merge_cells('B3:S3')
                sheet.unmerge_cells('B3:S3')
                sheet.merge_cells('B3:E3')

                sheet.merge_cells('I4:X4')
                sheet.unmerge_cells('I4:X4')
                sheet.merge_cells('I4:O4')

                sheet.merge_cells('I5:X5')
                sheet.unmerge_cells('I5:X5')
                sheet.merge_cells('I5:O5')

                sheet.merge_cells('I6:X6')
                sheet.unmerge_cells('I6:X6')
                sheet.merge_cells('I6:O6')

                sheet.merge_cells('I7:X7')
                sheet.unmerge_cells('I7:X7')
                sheet.merge_cells('I7:O7')

                sheet.merge_cells('I9:X9')
                sheet.unmerge_cells('I9:X9')
                sheet.merge_cells('I9:O9')

                sheet.merge_cells('B10:Y10')
                sheet.unmerge_cells('B10:Y10')
                sheet.merge_cells('B10:O10')


                sheet.merge_cells('P1:Y26')
                sheet.unmerge_cells('P1:Y26')

                sheet.merge_cells('Q13:R13')
                sheet.unmerge_cells('Q13:R13')

                sheet.merge_cells('S13:T13')
                sheet.unmerge_cells('S13:T13')

                amount = "J"+str(maxrow)+":Q"+str(maxrow)
                amount1 ="J"+str(maxrow)+":O"+str(maxrow)

                sheet.merge_cells(amount)
                sheet.unmerge_cells(amount)
                sheet.merge_cells(amount1)

                sheet.delete_cols(16, 25)

                print(rupees)

                #sheet['G33'].value = "=SpellNumber("+str(rupees)+")"
                print("O"+str(maxrow+2))
                sheet["E"+str(maxrow+2)].value = "DATE"
                sheet["O"+str(maxrow+2)].value="SIGNATURE"
                sheet['B13'].value = "SL"
                j=1
                for i in range(14, maxrow-1):
                    lll = "B" + str(i)
                # gg = int(maxrow/2)
                    #gg = gg+1

                    lllk = "B" + str(i)
                    sheet[lllk].value = j

                    j=j+1
                sheet['B13'].alignment = Alignment(horizontal='center', vertical='center')
                sheet[lll].alignment = Alignment(horizontal='center', vertical='center')

                print(maxrow)




                refernce_no = sheet['D14'].value

                #sheet.page_setup.fitTopage = 1
                # sheet.page_setup.fittowidth = 1
                #sheet.page_setup.fittoHeight = 0
                sheet.page_setup.scale = 72
                

                file.save("C:\\Users\\MY BOOK\\Desktop\\{}.xlsx".format(sheet['D14'].value))


            return render(request,'convert.html',{'message':"Congratz!!!The file(s) is Successfully Edited :)"})
        
        
    return render(request,'convert.html ',{'message':"Please Choose File(s) :)"})


def view(request):
    return render(request,'view.html')
    